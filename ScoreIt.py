import re
import shutil
from openpyxl import load_workbook

__author__ = 'FiftysixTimes7(PTJ)'
__version__ = 'Release 1.0'

if __name__ == '__main__':
    # Input file.
    while True:
        try:
            path = input('输入文件路径：').strip().strip('"\'')
            wb = load_workbook(path)
            wb.guess_types = True
            shutil.copyfile(path, path + '.bak')
            print('原文件已备份至 ' + path + '.bak')
        except FileNotFoundError as e:
            print(e)
            print('文件不存在')
        except PermissionError as e:
            print(e)
            print('权限不足')
        except OSError as e:
            print(e)
            print('路径错误')
        else:
            break

    # Choose sheet.
    while True:
        try:
            ws = wb[input('选择一个表格{}：'.format(wb.sheetnames))]
        except KeyError as e:
            print(e)
            print('表格不存在')
        else:
            break

    # Find ordinal.
    for column in ws.columns:
        for cell in column:
            if cell.value == '序号' or cell.value == '学号':
                order = cell
                break
        else:
            continue
        break
    else:
        raise ValueError('未找到“序号”或“学号”，请修改文件后重试')

    # Assign ordinal and name with row name.
    objects = {}
    for cell in ws[order.column]:
        if isinstance(cell.value, int):
            objects[str(cell.value)] = {'name': cell.offset(
                column=1).value, 'row': str(cell.row)}
            print('已导入数据：{}号 {}'.format(cell.value,
                                        objects[str(cell.value)]['name']))

    # Choose header.
    save = True

    def choose_header():
        headers = {}
        for cell_h in ws[order.row]:
            if cell_h.column > order.offset(column=1).column:
                headers[cell_h.value] = cell_h.column
        current = input(
            '请选择待输入列（输入一个不同于列表中的列来新建一列）{}：'.format(list(headers.keys())))
        if current not in headers:
            global save
            save = False
            choice = ws[order.row][-1].offset(column=1)
            choice.value = current
            headers[current] = choice.column
        current = headers[current]
        return current
    output = choose_header()

    # Main loop.
    while True:
        c = input(':')
        if c == 'q':
            if not save:
                c = input('是否保存？（y/n）')
                if c == 'y':
                    wb.save(path)
                    save = True
                    break
                elif c == 'n':
                    break
                else:
                    print('未知选项')
                    continue
            else:
                break
        elif c == 'c':
            output = choose_header()
        elif c == 's':
            wb.save(path)
            save = True
        elif c == 'p':
            for k in objects:
                print('{}号 {}：{}'.format(k, objects[k]['name'],
                                         ws[output + objects[k]['row']].value))
        elif re.match(r'[^,\s]+(,[^,\s]+)* \S+', c):
            numbers = c.split()[0].split(',')
            value = c.split()[1]
            for n in numbers:
                if not objects.get(n):
                    for k in objects:
                        if objects[k]['name'] == n:
                            n = k
                            break
                    else:
                        print('无法查询')
                        break
                ws[output + objects[n]['row']].value = value
                print('已导入数据：{}号 {}：{}'.format(n, objects[n]['name'], value))
            save = False
        else:
            print('格式错误')
